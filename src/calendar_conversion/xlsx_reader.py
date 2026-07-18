"""Read normalized calendar events from XLSX workbooks."""

from __future__ import annotations

import csv
from datetime import date, datetime, time
from io import StringIO
from os import PathLike
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .csv_reader import CSVReadError, _read_events_csv_with_source_rows
from .event import Event


class XLSXReadError(ValueError):
    """An error in the structure or contents of an events worksheet."""

    def __init__(
        self,
        message: str,
        *,
        sheet_name: str | None = None,
        row_number: int | None = None,
    ) -> None:
        self.sheet_name = sheet_name
        self.row_number = row_number

        location_parts = []
        if sheet_name is not None:
            location_parts.append(f"sheet {sheet_name!r}")
        if row_number is not None:
            location_parts.append(f"row {row_number}")
        prefix = ", ".join(location_parts)
        super().__init__(f"{prefix}: {message}" if prefix else message)


def read_events_xlsx(
    source: str | PathLike[str] | BinaryIO,
    *,
    sheet_name: str | None = None,
) -> list[Event]:
    """Read events from one worksheet in an XLSX path or binary stream.

    The active worksheet is read by default. Pass ``sheet_name`` to select a
    worksheet explicitly. Completely empty data rows are ignored.
    """
    events, _, _ = _read_events_xlsx_with_source_rows(
        source, sheet_name=sheet_name
    )
    return events


def _read_events_xlsx_with_source_rows(
    source: str | PathLike[str] | BinaryIO,
    *,
    sheet_name: str | None = None,
) -> tuple[list[Event], list[int], str]:
    """Read events and retain their original worksheet row numbers."""
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise XLSXReadError(f"could not read XLSX workbook: {error}") from error

    try:
        if sheet_name is None:
            worksheet = workbook.active
        else:
            if sheet_name not in workbook.sheetnames:
                raise XLSXReadError(
                    f"worksheet does not exist; available worksheets: "
                    f"{', '.join(workbook.sheetnames)}",
                    sheet_name=sheet_name,
                )
            worksheet = workbook[sheet_name]

        csv_source, source_rows = _worksheet_as_csv(worksheet)
        try:
            events, normalized_rows = _read_events_csv_with_source_rows(
                csv_source
            )
        except CSVReadError as error:
            row_number = None
            if error.row_number is not None:
                row_number = source_rows[error.row_number - 1]
            raise XLSXReadError(
                error.message,
                sheet_name=worksheet.title,
                row_number=row_number,
            ) from error
        event_source_rows = [source_rows[row - 1] for row in normalized_rows]
        return events, event_source_rows, worksheet.title
    finally:
        workbook.close()


def _worksheet_as_csv(worksheet: object) -> tuple[StringIO, list[int]]:
    """Normalize worksheet values into the CSV reader's input contract."""
    output = StringIO(newline="")
    writer = csv.writer(output)
    source_rows: list[int] = []

    rows = worksheet.iter_rows(values_only=True)  # type: ignore[attr-defined]
    header = next(rows, None)
    if header is None:
        output.seek(0)
        return output, source_rows

    header_values = [_string_value(value) for value in header]
    writer.writerow(header_values)
    source_rows.append(1)

    for row_number, row in enumerate(rows, start=2):
        if all(_is_empty(value) for value in row):
            continue

        writer.writerow(
            _normalized_value(value, header_values[column_index])
            for column_index, value in enumerate(row)
        )
        source_rows.append(row_number)

    output.seek(0)
    return output, source_rows


def _normalized_value(value: object, column: str) -> str:
    """Convert Excel-native values to the normalized CSV representation."""
    normalized_column = column.strip()

    if value is None:
        return ""
    if normalized_column == "all_date" and isinstance(value, bool):
        return "true" if value else "false"
    if normalized_column in {"start_date", "end_date"}:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
    if normalized_column in {"start_time", "end_time"}:
        if isinstance(value, datetime):
            return value.time().isoformat()
        if isinstance(value, time):
            return value.isoformat()

    return _string_value(value)


def _string_value(value: object) -> str:
    return "" if value is None else str(value)


def _is_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
