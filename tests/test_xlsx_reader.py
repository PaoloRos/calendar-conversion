"""Tests for the normalized events XLSX reader."""

from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from calendar_conversion import XLSXReadError, read_events_xlsx


HEADER = [
    "id",
    "summary",
    "all_date",
    "start_date",
    "start_time",
    "end_date",
    "end_time",
    "location",
    "description",
]


def workbook_bytes(rows: list[list[object]], *, title: str = "Events") -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


class XLSXReaderTests(unittest.TestCase):
    def test_reads_excel_dates_times_booleans_and_skips_empty_rows(self) -> None:
        source = workbook_bytes(
            [
                HEADER,
                [
                    "shift-1",
                    "Station shift",
                    False,
                    date(2026, 7, 10),
                    time(8, 0),
                    datetime(2026, 7, 10),
                    time(20, 0),
                    "Central station",
                    "Day team",
                ],
                [None] * len(HEADER),
                [
                    "training-1",
                    "Training day",
                    True,
                    date(2026, 7, 11),
                    None,
                    date(2026, 7, 11),
                    None,
                    None,
                    None,
                ],
            ]
        )

        events = read_events_xlsx(source)

        self.assertEqual(len(events), 2)
        self.assertEqual(events[0].start_date, date(2026, 7, 10))
        self.assertEqual(events[0].start_time, time(8, 0))
        self.assertFalse(events[0].all_day)
        self.assertTrue(events[1].all_day)
        self.assertEqual(events[1].location, "")

    def test_reads_path_and_named_worksheet(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Notes"
        worksheet = workbook.create_sheet("Schedule")
        worksheet.append(HEADER[:-2])
        worksheet.append(
            [
                "event-1",
                "Event",
                "false",
                "2026-07-10",
                "08:00",
                "2026-07-10",
                "09:00",
            ]
        )

        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory, "events.xlsx")
            workbook.save(path)
            workbook.close()

            event = read_events_xlsx(path, sheet_name="Schedule")[0]

        self.assertEqual(event.id, "event-1")
        self.assertEqual(event.location, "")

    def test_reports_the_original_worksheet_row_after_empty_rows(self) -> None:
        source = workbook_bytes(
            [
                HEADER,
                [None] * len(HEADER),
                [
                    "event-1",
                    "Event",
                    "perhaps",
                    "2026-07-10",
                    None,
                    "2026-07-10",
                    None,
                    None,
                    None,
                ],
            ]
        )

        with self.assertRaisesRegex(
            XLSXReadError,
            r"sheet 'Events', row 3: all_date",
        ) as raised:
            read_events_xlsx(source)

        self.assertEqual(raised.exception.sheet_name, "Events")
        self.assertEqual(raised.exception.row_number, 3)

    def test_rejects_missing_named_worksheet(self) -> None:
        source = workbook_bytes([HEADER])

        with self.assertRaisesRegex(XLSXReadError, "worksheet does not exist"):
            read_events_xlsx(source, sheet_name="Missing")

    def test_rejects_missing_required_columns(self) -> None:
        source = workbook_bytes([["id", "summary"], ["event-1", "Event"]])

        with self.assertRaisesRegex(XLSXReadError, "missing required column"):
            read_events_xlsx(source)

    def test_wraps_invalid_workbook_errors(self) -> None:
        with self.assertRaisesRegex(XLSXReadError, "could not read XLSX"):
            read_events_xlsx(BytesIO(b"not an XLSX workbook"))


if __name__ == "__main__":
    unittest.main()
